#!/usr/bin/env python3
"""
Bay Area IT/CS Consultancy Contact Finder
==========================================
This script uses Apollo.io and/or Hunter.io APIs to find people who recruit
for IT / Computer Science roles at top Bay Area consultancies.

Key features:
  - MULTIPLE API keys per service with auto-failover on rate limit
  - INCREMENTAL: re-running the script skips people and searches already
    completed in a previous run, so you never waste credits on duplicates.
    Just add new keys to .env and re-run to continue where you left off.

Setup:
  1. Create free accounts on Apollo.io and/or Hunter.io
     (create as many as you want for more quota)
  2. Copy your API key(s) from their dashboards
  3. Create a .env file in this directory (see .env.example)

.env file format (multiple keys supported):
  APOLLO_API_KEY_1=key_one
  APOLLO_API_KEY_2=key_two
  HUNTER_API_KEY_1=key_one
  HUNTER_API_KEY_2=key_two
  # ... add as many as you like (_3, _4, etc.)

  # Single-key shortcuts also work:
  APOLLO_API_KEY=single_key
  HUNTER_API_KEY=single_key

Usage:
  python find_consultancy_contacts.py                              # uses .env
  python find_consultancy_contacts.py --apollo-keys K1 K2 K3      # multiple Apollo keys
  python find_consultancy_contacts.py --hunter-keys K1 K2          # multiple Hunter keys
  python find_consultancy_contacts.py --no-enrich                  # skip Apollo enrichment
  python find_consultancy_contacts.py --fresh                      # ignore cache, start over
  python find_consultancy_contacts.py --help

Output:
  Bay_Area_IT_Consultancies_Contacts.xlsx  — the contact spreadsheet
  .search_cache.json                       — tracks completed API calls
"""

import argparse
import json
import os
import sys
import time
from pathlib import Path

try:
    import requests
except ImportError:
    print("ERROR: 'requests' library is required. Install it with:")
    print("  pip install requests")
    sys.exit(1)

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("ERROR: 'openpyxl' library is required. Install it with:")
    print("  pip install openpyxl")
    sys.exit(1)


# ──────────────────────────────────────────────
# Configuration
# ──────────────────────────────────────────────

CONSULTANCIES = [
    {
        "name": "Accenture",
        "domain": "accenture.com",
        "careers_url": "https://www.accenture.com/us-en/careers",
    },
    {
        "name": "Deloitte",
        "domain": "deloitte.com",
        "careers_url": "https://www2.deloitte.com/us/en/careers.html",
    },
    {
        "name": "McKinsey & Company",
        "domain": "mckinsey.com",
        "careers_url": "https://www.mckinsey.com/careers",
    },
    {
        "name": "Slalom Consulting",
        "domain": "slalom.com",
        "careers_url": "https://www.slalom.com/us/en/careers",
    },
    {
        "name": "Cognizant",
        "domain": "cognizant.com",
        "careers_url": "https://careers.cognizant.com",
    },
    # ── Batch 2 ──
    {
        "name": "Capgemini",
        "domain": "capgemini.com",
        "careers_url": "https://www.capgemini.com/careers/",
    },
    {
        "name": "Infosys",
        "domain": "infosys.com",
        "careers_url": "https://www.infosys.com/careers/",
    },
    {
        "name": "KPMG",
        "domain": "kpmg.com",
        "careers_url": "https://kpmg.com/us/en/careers.html",
    },
    {
        "name": "EY (Ernst & Young)",
        "domain": "ey.com",
        "careers_url": "https://www.ey.com/en_us/careers",
    },
    {
        "name": "Wipro",
        "domain": "wipro.com",
        "careers_url": "https://careers.wipro.com/",
    },
    # ── Batch 3 ──
    {
        "name": "TCS (Tata Consultancy Services)",
        "domain": "tcs.com",
        "careers_url": "https://www.tcs.com/careers",
    },
    {
        "name": "PwC",
        "domain": "pwc.com",
        "careers_url": "https://www.pwc.com/us/en/careers.html",
    },
    {
        "name": "Booz Allen Hamilton",
        "domain": "bah.com",
        "careers_url": "https://www.boozallen.com/careers.html",
    },
    {
        "name": "HCL Technologies",
        "domain": "hcltech.com",
        "careers_url": "https://www.hcltech.com/careers",
    },
    {
        "name": "West Monroe",
        "domain": "westmonroe.com",
        "careers_url": "https://www.westmonroe.com/careers",
    },
    # ── Batch 4 (IT dept only — final 5 credits) ──
    {
        "name": "IBM Consulting",
        "domain": "ibm.com",
        "careers_url": "https://www.ibm.com/careers",
    },
    {
        "name": "Thoughtworks",
        "domain": "thoughtworks.com",
        "careers_url": "https://www.thoughtworks.com/careers",
    },
    {
        "name": "NTT DATA",
        "domain": "nttdata.com",
        "careers_url": "https://www.nttdata.com/global/en/careers",
    },
    {
        "name": "CGI Group",
        "domain": "cgi.com",
        "careers_url": "https://www.cgi.com/en/careers",
    },
    {
        "name": "Tech Mahindra",
        "domain": "techmahindra.com",
        "careers_url": "https://careers.techmahindra.com/",
    },
]

# Job titles to search for — people involved in IT/CS hiring
APOLLO_TITLES = [
    "IT Recruiter",
    "Technology Recruiter",
    "Technical Recruiter",
    "Talent Acquisition",
    "Recruiting Manager",
    "Hiring Manager IT",
    "Technology Hiring Manager",
    "Head of Recruiting",
    "Director of Talent Acquisition",
    "University Recruiter",
    "Campus Recruiter Technology",
]

APOLLO_LOCATIONS = [
    "San Francisco, California, United States",
    "San Jose, California, United States",
    "Palo Alto, California, United States",
    "Mountain View, California, United States",
    "Oakland, California, United States",
    "Bay Area",
]

# Rate limiting settings
RATE_LIMIT_DELAY = 1.0  # seconds between API calls


# ──────────────────────────────────────────────
# Utility: load .env file
# ──────────────────────────────────────────────

def load_env_file(env_path: str = ".env"):
    """Load key=value pairs from a .env file into environment variables."""
    path = Path(env_path)
    if not path.exists():
        return
    with open(path) as f:
        for line in f:
            line = line.strip()
            if not line or line.startswith("#"):
                continue
            if "=" in line:
                key, _, value = line.partition("=")
                value = value.strip().strip('"').strip("'")
                # Strip inline comments (e.g., "api_key_here # my_email@gmail.com")
                if " #" in value:
                    value = value[:value.index(" #")].strip()
                os.environ.setdefault(key.strip(), value)


def collect_keys_from_env(prefix: str) -> list:
    """
    Collect API keys from environment variables.
    Supports both numbered keys (PREFIX_1, PREFIX_2, ...) and a single
    key (PREFIX) as a fallback.
    """
    keys = []
    for i in range(1, 21):
        val = os.environ.get(f"{prefix}_{i}", "").strip()
        if val and val not in keys:
            keys.append(val)
    single = os.environ.get(prefix, "").strip()
    if single and single not in keys:
        keys.append(single)
    return keys


# ──────────────────────────────────────────────
# Search Cache — tracks completed API calls
# so re-runs don't repeat the same searches
# ──────────────────────────────────────────────

class SearchCache:
    """
    Persists which API calls have already been completed to a JSON file.
    This avoids re-running the same Hunter domain-search or Apollo search
    on subsequent runs, saving API credits.

    Cache keys look like:
      "hunter|accenture.com|it"
      "hunter|accenture.com|hr"
      "apollo_search|accenture.com|titles"
      "apollo_search|accenture.com|broad"
    """

    def __init__(self, cache_path: str = ".search_cache.json"):
        self.cache_path = Path(cache_path)
        self._completed: set = set()
        self._load()

    def _load(self):
        if self.cache_path.exists():
            try:
                with open(self.cache_path) as f:
                    data = json.load(f)
                self._completed = set(data.get("completed_searches", []))
            except (json.JSONDecodeError, KeyError):
                self._completed = set()

    def save(self):
        with open(self.cache_path, "w") as f:
            json.dump({"completed_searches": sorted(self._completed)}, f, indent=2)

    def is_done(self, key: str) -> bool:
        return key in self._completed

    def mark_done(self, key: str):
        self._completed.add(key)
        self.save()  # persist immediately so progress isn't lost on crash

    def clear(self):
        self._completed = set()
        if self.cache_path.exists():
            self.cache_path.unlink()

    @property
    def count(self) -> int:
        return len(self._completed)

    def __repr__(self):
        return f"SearchCache({self.count} completed searches)"


# ──────────────────────────────────────────────
# Existing Excel loader — reads contacts from a
# previous run so we don't re-process them
# ──────────────────────────────────────────────

def load_existing_contacts(excel_path: str) -> tuple:
    """
    Load contacts from a previously generated Excel file.

    Returns:
        (contacts_list, seen_set, enriched_set)
        - contacts_list: list of contact dicts (same format as collect_contacts)
        - seen_set: set of dedup keys (name|company and email keys)
        - enriched_set: set of name|company keys that already have an email
                        (so we skip enrichment for them)
    """
    path = Path(excel_path)
    if not path.exists():
        return [], set(), set()

    try:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    except Exception as e:
        print(f"  [Cache] Could not read existing Excel: {e}")
        return [], set(), set()

    # Find the contacts sheet
    sheet_name = "IT CS Recruiting Contacts"
    if sheet_name not in wb.sheetnames:
        wb.close()
        return [], set(), set()

    ws = wb[sheet_name]
    contacts = []
    seen = set()
    enriched = set()

    # Expected columns (1-indexed): Consultancy, First Name, Last Name, Email,
    # Job Title, LinkedIn Profile URL, Phone, Source, Notes
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
        if not row or not row[0]:  # skip empty rows
            continue

        # Unpack — handle rows that might have fewer columns
        vals = list(row) + [None] * 9  # pad to ensure we have enough
        consultancy = str(vals[0] or "").strip()
        first_name  = str(vals[1] or "").strip()
        last_name   = str(vals[2] or "").strip()
        email       = str(vals[3] or "").strip()
        job_title   = str(vals[4] or "").strip()
        linkedin    = str(vals[5] or "").strip()
        phone       = str(vals[6] or "").strip()
        source      = str(vals[7] or "").strip()
        notes       = str(vals[8] or "").strip()

        if not consultancy or (not first_name and not last_name and not email):
            continue

        contact = {
            "consultancy": consultancy,
            "first_name": first_name,
            "last_name": last_name,
            "email": email,
            "job_title": job_title,
            "linkedin_url": linkedin,
            "phone": phone,
            "source": source,
            "notes": notes,
        }
        contacts.append(contact)

        # Build dedup keys
        name_key = f"{first_name}|{last_name}|{consultancy}".lower()
        seen.add(name_key)
        if email:
            seen.add(email.lower())
            enriched.add(name_key)  # this person already has an email

    wb.close()
    return contacts, seen, enriched


# ──────────────────────────────────────────────
# Key Pool — manages multiple API keys with
# automatic failover on rate limit / exhaustion
# ──────────────────────────────────────────────

class KeyPool:
    """
    Manages a pool of API keys for a single service.
    When a key is exhausted (rate limited / quota used up), it is
    moved to the "exhausted" set and the next key becomes active.
    """

    def __init__(self, service_name: str, keys: list):
        self.service_name = service_name
        self.keys = list(keys)
        self.exhausted = set()
        self._current_idx = 0

        if not self.keys:
            raise ValueError(f"No API keys provided for {service_name}")

    @property
    def total(self) -> int:
        return len(self.keys)

    @property
    def active_count(self) -> int:
        return self.total - len(self.exhausted)

    @property
    def current_key(self) -> str | None:
        if self._current_idx in self.exhausted:
            if not self._rotate():
                return None
        return self.keys[self._current_idx]

    @property
    def current_label(self) -> str:
        return f"key #{self._current_idx + 1} of {self.total}"

    @property
    def all_exhausted(self) -> bool:
        return len(self.exhausted) >= self.total

    def mark_exhausted(self) -> bool:
        self.exhausted.add(self._current_idx)
        key_num = self._current_idx + 1
        print(f"  [{self.service_name}] Key #{key_num} exhausted "
              f"({self.active_count} of {self.total} keys remaining)")
        if self.all_exhausted:
            print(f"  [{self.service_name}] ALL {self.total} keys exhausted! "
                  f"Cannot make further requests.")
            return False
        return self._rotate()

    def _rotate(self) -> bool:
        for _ in range(self.total):
            self._current_idx = (self._current_idx + 1) % self.total
            if self._current_idx not in self.exhausted:
                key_num = self._current_idx + 1
                print(f"  [{self.service_name}] Switched to key #{key_num} of {self.total}")
                return True
        return False

    def __repr__(self):
        return (f"KeyPool({self.service_name}: {self.active_count}/{self.total} active, "
                f"current=#{self._current_idx + 1})")


# ──────────────────────────────────────────────
# Apollo.io API Client (multi-key)
# ──────────────────────────────────────────────

class ApolloClient:
    BASE_URL = "https://api.apollo.io/api/v1"
    EXHAUSTION_CODES = {429, 403}

    def __init__(self, key_pool: KeyPool):
        self.pool = key_pool

    def _headers(self) -> dict:
        return {
            "Content-Type": "application/json",
            "Cache-Control": "no-cache",
            "X-Api-Key": self.pool.current_key,
        }

    def _request_with_failover(self, method: str, url: str, **kwargs) -> dict:
        while not self.pool.all_exhausted:
            key = self.pool.current_key
            if key is None:
                return {}
            kwargs["headers"] = self._headers()
            kwargs.setdefault("timeout", 30)
            try:
                resp = requests.request(method, url, **kwargs)
                if resp.status_code in self.EXHAUSTION_CODES:
                    print(f"  [Apollo] HTTP {resp.status_code} — "
                          f"rate limit / quota hit on {self.pool.current_label}")
                    if not self.pool.mark_exhausted():
                        return {}
                    time.sleep(1)
                    continue
                if resp.status_code == 401:
                    print(f"  [Apollo] HTTP 401 — invalid key ({self.pool.current_label}), skipping it")
                    if not self.pool.mark_exhausted():
                        return {}
                    continue
                resp.raise_for_status()
                return resp.json()
            except requests.exceptions.HTTPError as e:
                print(f"  [Apollo] HTTP error: {e}")
                return {}
            except requests.exceptions.RequestException as e:
                print(f"  [Apollo] Request error: {e}")
                return {}
        return {}

    def search_people(self, titles: list, locations: list, org_domain: str = None,
                      per_page: int = 25, page: int = 1) -> dict:
        url = f"{self.BASE_URL}/mixed_people/api_search"
        payload = {
            "person_titles": titles,
            "person_locations": locations,
            "per_page": per_page,
            "page": page,
        }
        if org_domain:
            payload["q_organization_domains"] = org_domain
        return self._request_with_failover("POST", url, json=payload)

    def enrich_person(self, first_name: str, last_name: str, domain: str,
                      reveal_email: bool = True) -> dict:
        url = f"{self.BASE_URL}/people/match"
        payload = {
            "first_name": first_name,
            "last_name": last_name,
            "organization_domain": domain,
            "reveal_personal_emails": reveal_email,
        }
        return self._request_with_failover("POST", url, json=payload)


# ──────────────────────────────────────────────
# Hunter.io API Client (multi-key)
# ──────────────────────────────────────────────

class HunterClient:
    BASE_URL = "https://api.hunter.io/v2"
    EXHAUSTION_CODES = {429, 403}

    def __init__(self, key_pool: KeyPool):
        self.pool = key_pool

    def _request_with_failover(self, url: str, params: dict) -> dict:
        while not self.pool.all_exhausted:
            key = self.pool.current_key
            if key is None:
                return {}
            params["api_key"] = key
            try:
                resp = requests.get(url, params=params, timeout=30)
                if resp.status_code in self.EXHAUSTION_CODES:
                    print(f"  [Hunter] HTTP {resp.status_code} — "
                          f"rate limit / quota hit on {self.pool.current_label}")
                    if not self.pool.mark_exhausted():
                        return {}
                    time.sleep(1)
                    continue
                if resp.status_code == 401:
                    print(f"  [Hunter] HTTP 401 — invalid key ({self.pool.current_label}), skipping it")
                    if not self.pool.mark_exhausted():
                        return {}
                    continue
                resp.raise_for_status()
                return resp.json()
            except requests.exceptions.HTTPError as e:
                print(f"  [Hunter] HTTP error: {e}")
                return {}
            except requests.exceptions.RequestException as e:
                print(f"  [Hunter] Request error: {e}")
                return {}
        return {}

    def domain_search(self, domain: str, department: str = "it",
                      seniority: str = None, limit: int = 10,
                      offset: int = 0) -> dict:
        params = {
            "domain": domain,
            "department": department,
            "limit": limit,
            "offset": offset,
            "type": "personal",
            "required_field": "full_name",
        }
        if seniority:
            params["seniority"] = seniority
        return self._request_with_failover(f"{self.BASE_URL}/domain-search", params)

    def domain_search_recruiting(self, domain: str, limit: int = 10) -> dict:
        params = {
            "domain": domain,
            "department": "hr",
            "limit": limit,
            "type": "personal",
            "required_field": "full_name",
        }
        return self._request_with_failover(f"{self.BASE_URL}/domain-search", params)


# ──────────────────────────────────────────────
# Contact collector (incremental)
# ──────────────────────────────────────────────

def collect_contacts(
    apollo_pool: KeyPool = None,
    hunter_pool: KeyPool = None,
    enrich_apollo: bool = True,
    existing_contacts: list = None,
    seen: set = None,
    enriched: set = None,
    cache: SearchCache = None,
    hunter_it_only: bool = False,
) -> list:
    """
    Collect contacts from both APIs for all consultancies.
    Merges with existing_contacts and skips already-seen people and
    already-completed searches.

    If hunter_it_only=True, only search the IT department on Hunter
    (saves credits — 1 per company instead of 3).
    """
    # Start from existing data
    contacts = list(existing_contacts or [])
    seen = set(seen or set())
    enriched = set(enriched or set())
    cache = cache or SearchCache()

    new_contacts = 0
    skipped_searches = 0
    skipped_enrichments = 0

    apollo = ApolloClient(apollo_pool) if apollo_pool else None
    hunter = HunterClient(hunter_pool) if hunter_pool else None

    for firm in CONSULTANCIES:
        name = firm["name"]
        domain = firm["domain"]
        print(f"\n{'='*60}")
        print(f"Searching: {name} ({domain})")
        print(f"{'='*60}")

        # ── Apollo.io search ──
        if apollo and not apollo.pool.all_exhausted:

            # --- Narrow title search ---
            cache_key_titles = f"apollo_search|{domain}|titles"
            if cache.is_done(cache_key_titles):
                print(f"\n  [Apollo] Title search at {name}: CACHED (skipping)")
                skipped_searches += 1
            else:
                print(f"\n  [Apollo] Searching for IT/recruiting titles at {name}... "
                      f"(using {apollo.pool.current_label})")
                result = apollo.search_people(
                    titles=APOLLO_TITLES,
                    locations=APOLLO_LOCATIONS,
                    org_domain=domain,
                    per_page=10,
                )
                people = result.get("people", [])
                print(f"  [Apollo] Found {len(people)} people")

                for person in people:
                    first = person.get("first_name", "")
                    last = person.get("last_name", "")
                    title = person.get("title", "")
                    linkedin = person.get("linkedin_url", "")
                    email = person.get("email", "")

                    name_key = f"{first}|{last}|{name}".lower()
                    if name_key in seen:
                        continue
                    seen.add(name_key)

                    # Enrichment: skip if already enriched or all keys exhausted
                    if enrich_apollo and not email and first and last and not apollo.pool.all_exhausted:
                        if name_key in enriched:
                            print(f"    Skipping enrichment (already has email): {first} {last}")
                            skipped_enrichments += 1
                        else:
                            print(f"    Enriching: {first} {last}... ({apollo.pool.current_label})")
                            enrich_result = apollo.enrich_person(first, last, domain)
                            person_data = enrich_result.get("person", {})
                            if person_data:
                                email = person_data.get("email", "")
                                if not linkedin:
                                    linkedin = person_data.get("linkedin_url", "")
                                if email:
                                    enriched.add(name_key)
                            time.sleep(RATE_LIMIT_DELAY)

                    contacts.append({
                        "consultancy": name,
                        "first_name": first,
                        "last_name": last,
                        "email": email or "",
                        "job_title": title,
                        "linkedin_url": linkedin or "",
                        "phone": "",
                        "source": "Apollo.io",
                        "notes": "",
                    })
                    new_contacts += 1

                # Mark search as completed (even if 0 results — it was done)
                if not apollo.pool.all_exhausted or people:
                    cache.mark_done(cache_key_titles)

                time.sleep(RATE_LIMIT_DELAY)

            # --- Broader title search ---
            cache_key_broad = f"apollo_search|{domain}|broad"
            if cache.is_done(cache_key_broad):
                print(f"  [Apollo] Broad search at {name}: CACHED (skipping)")
                skipped_searches += 1
            elif not apollo.pool.all_exhausted:
                print(f"  [Apollo] Broader search for talent acquisition at {name}...")
                broader_result = apollo.search_people(
                    titles=["Talent Acquisition", "Recruiter", "Staffing"],
                    locations=APOLLO_LOCATIONS,
                    org_domain=domain,
                    per_page=5,
                )
                for person in broader_result.get("people", []):
                    first = person.get("first_name", "")
                    last = person.get("last_name", "")
                    name_key = f"{first}|{last}|{name}".lower()
                    if name_key in seen:
                        continue
                    seen.add(name_key)

                    title = person.get("title", "")
                    linkedin = person.get("linkedin_url", "")
                    email = person.get("email", "")

                    if enrich_apollo and not email and first and last and not apollo.pool.all_exhausted:
                        if name_key in enriched:
                            print(f"    Skipping enrichment (already has email): {first} {last}")
                            skipped_enrichments += 1
                        else:
                            print(f"    Enriching: {first} {last}... ({apollo.pool.current_label})")
                            enrich_result = apollo.enrich_person(first, last, domain)
                            person_data = enrich_result.get("person", {})
                            if person_data:
                                email = person_data.get("email", "")
                                if not linkedin:
                                    linkedin = person_data.get("linkedin_url", "")
                                if email:
                                    enriched.add(name_key)
                            time.sleep(RATE_LIMIT_DELAY)

                    contacts.append({
                        "consultancy": name,
                        "first_name": first,
                        "last_name": last,
                        "email": email or "",
                        "job_title": title,
                        "linkedin_url": linkedin or "",
                        "phone": "",
                        "source": "Apollo.io",
                        "notes": "",
                    })
                    new_contacts += 1

                if not apollo.pool.all_exhausted:
                    cache.mark_done(cache_key_broad)
                time.sleep(RATE_LIMIT_DELAY)

        elif apollo and apollo.pool.all_exhausted:
            print(f"\n  [Apollo] Skipping {name} — all Apollo keys exhausted")

        # ── Hunter.io search ──
        if hunter and not hunter.pool.all_exhausted:

            # Helper to process Hunter results
            def process_hunter_emails(emails_list, dept_label):
                nonlocal new_contacts
                for entry in emails_list:
                    email = entry.get("value", "")
                    first = entry.get("first_name", "")
                    last = entry.get("last_name", "")
                    title = entry.get("position", "")
                    linkedin = entry.get("linkedin", "")
                    phone = entry.get("phone_number", "")
                    confidence = entry.get("confidence", 0)

                    dedup_key_email = email.lower() if email else ""
                    name_key = f"{first}|{last}|{name}".lower()
                    if dedup_key_email in seen or name_key in seen:
                        continue
                    if dedup_key_email:
                        seen.add(dedup_key_email)
                    seen.add(name_key)
                    if email:
                        enriched.add(name_key)

                    notes = f"Confidence: {confidence}%"
                    if dept_label:
                        notes += f"; Dept: {dept_label}"

                    contacts.append({
                        "consultancy": name,
                        "first_name": first or "",
                        "last_name": last or "",
                        "email": email,
                        "job_title": title or "",
                        "linkedin_url": linkedin or "",
                        "phone": phone or "",
                        "source": "Hunter.io",
                        "notes": notes,
                    })
                    new_contacts += 1

            # --- IT department ---
            cache_key_it = f"hunter|{domain}|it"
            if cache.is_done(cache_key_it):
                print(f"\n  [Hunter] IT search at {domain}: CACHED (skipping)")
                skipped_searches += 1
            else:
                print(f"\n  [Hunter] Searching IT department at {domain}... "
                      f"(using {hunter.pool.current_label})")
                it_result = hunter.domain_search(domain, department="it", limit=10)
                it_emails = it_result.get("data", {}).get("emails", [])
                print(f"  [Hunter] Found {len(it_emails)} IT contacts")
                process_hunter_emails(it_emails, "IT")
                if not hunter.pool.all_exhausted or it_emails:
                    cache.mark_done(cache_key_it)
                time.sleep(RATE_LIMIT_DELAY)

            # --- HR / Recruiting department ---
            if not hunter_it_only:
                cache_key_hr = f"hunter|{domain}|hr"
                if cache.is_done(cache_key_hr):
                    print(f"  [Hunter] HR/recruiting search at {domain}: CACHED (skipping)")
                    skipped_searches += 1
                elif not hunter.pool.all_exhausted:
                    print(f"  [Hunter] Searching HR/recruiting department at {domain}... "
                          f"(using {hunter.pool.current_label})")
                    hr_result = hunter.domain_search_recruiting(domain, limit=10)
                    hr_emails = hr_result.get("data", {}).get("emails", [])
                    print(f"  [Hunter] Found {len(hr_emails)} HR/recruiting contacts")
                    process_hunter_emails(hr_emails, "HR/Recruiting")
                    if not hunter.pool.all_exhausted or hr_emails:
                        cache.mark_done(cache_key_hr)
                    time.sleep(RATE_LIMIT_DELAY)

            # --- Management ---
            if not hunter_it_only:
                cache_key_mgmt = f"hunter|{domain}|management"
                if cache.is_done(cache_key_mgmt):
                    print(f"  [Hunter] Management search at {domain}: CACHED (skipping)")
                    skipped_searches += 1
                elif not hunter.pool.all_exhausted:
                    print(f"  [Hunter] Searching management at {domain}... "
                          f"(using {hunter.pool.current_label})")
                    mgmt_result = hunter.domain_search(domain, department="management",
                                                        seniority="senior,executive", limit=5)
                    mgmt_emails = mgmt_result.get("data", {}).get("emails", [])
                    print(f"  [Hunter] Found {len(mgmt_emails)} management contacts")
                    process_hunter_emails(mgmt_emails, "Management")
                    if not hunter.pool.all_exhausted or mgmt_emails:
                        cache.mark_done(cache_key_mgmt)
                    time.sleep(RATE_LIMIT_DELAY)

        elif hunter and hunter.pool.all_exhausted:
            print(f"\n  [Hunter] Skipping {name} — all Hunter keys exhausted")

    print(f"\n  --- Incremental stats ---")
    print(f"  New contacts added this run:  {new_contacts}")
    print(f"  Searches skipped (cached):    {skipped_searches}")
    print(f"  Enrichments skipped (cached): {skipped_enrichments}")

    return contacts


# ──────────────────────────────────────────────
# Excel writer
# ──────────────────────────────────────────────

def write_excel(contacts: list, output_path: str = "Bay_Area_IT_Consultancies_Contacts.xlsx"):
    """Write contacts to a formatted Excel file."""
    wb = openpyxl.Workbook()

    # Styles
    header_font = Font(name="Calibri", bold=True, size=12, color="FFFFFF")
    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    normal_font = Font(name="Calibri", size=11)
    green_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    def style_header(ws, headers, col_widths):
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = thin_border
        for i, w in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions

    # ── Sheet 1: All Contacts ──
    ws1 = wb.active
    ws1.title = "IT CS Recruiting Contacts"
    headers = ["Consultancy", "First Name", "Last Name", "Email",
               "Job Title", "LinkedIn Profile URL", "Phone", "Source", "Notes"]
    col_widths = [22, 15, 15, 35, 35, 45, 16, 12, 40]
    style_header(ws1, headers, col_widths)

    for row_idx, c in enumerate(contacts, 2):
        values = [c["consultancy"], c["first_name"], c["last_name"], c["email"],
                  c["job_title"], c["linkedin_url"], c["phone"], c["source"], c["notes"]]
        for col_idx, val in enumerate(values, 1):
            cell = ws1.cell(row=row_idx, column=col_idx, value=val)
            cell.font = normal_font
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = thin_border
            if c["email"]:
                cell.fill = green_fill

    # ── Sheet 2: Consultancies Overview ──
    ws2 = wb.create_sheet("Consultancies Overview")
    headers2 = ["Consultancy", "Domain", "Careers URL",
                "Contacts Found", "With Email"]
    col_widths2 = [22, 22, 50, 16, 14]
    style_header(ws2, headers2, col_widths2)

    for row_idx, firm in enumerate(CONSULTANCIES, 2):
        firm_name = firm["name"]
        firm_contacts = [c for c in contacts if c["consultancy"] == firm_name]
        with_email = len([c for c in firm_contacts if c["email"]])
        values = [firm_name, firm["domain"], firm["careers_url"],
                  len(firm_contacts), with_email]
        for col_idx, val in enumerate(values, 1):
            cell = ws2.cell(row=row_idx, column=col_idx, value=val)
            cell.font = normal_font
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = thin_border

    # ── Sheet 3: Instructions ──
    ws3 = wb.create_sheet("README")
    instructions = [
        ["Item", "Details"],
        ["Generated by", "find_consultancy_contacts.py"],
        ["Target firms", ", ".join(f["name"] for f in CONSULTANCIES)],
        ["Search focus", "IT, Computer Science, Technology recruiting contacts"],
        ["Location filter", "San Francisco Bay Area, CA"],
        ["Green rows", "Contacts where an email address was found"],
        ["Multi-key", "Supports multiple API keys per service with auto-failover"],
        ["Incremental", "Re-running skips already-found contacts and completed searches"],
        ["Cache file", ".search_cache.json tracks completed API calls"],
        ["Fresh start", "Use --fresh to clear cache and start over"],
        ["Apollo credits", "~1 credit per person enriched (search is free)"],
        ["Hunter credits", "1 request per domain-search call (free plan: 25 searches/mo)"],
        ["Tip", "Add more API keys in .env to increase your total quota"],
    ]
    style_header(ws3, ["Item", "Details"], [25, 80])
    for row_idx, row in enumerate(instructions[1:], 2):
        for col_idx, val in enumerate(row, 1):
            cell = ws3.cell(row=row_idx, column=col_idx, value=val)
            cell.font = normal_font
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = thin_border

    wb.save(output_path)
    print(f"\nExcel file saved: {output_path}")
    return output_path


# ──────────────────────────────────────────────
# Main
# ──────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(
        description="Find IT/CS recruiting contacts at top Bay Area consultancies",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Multiple API keys (auto-failover when one is exhausted):
  In .env file:
    APOLLO_API_KEY_1=key_one
    APOLLO_API_KEY_2=key_two
    HUNTER_API_KEY_1=key_one
    HUNTER_API_KEY_2=key_two

  Or via command line:
    python find_consultancy_contacts.py --apollo-keys K1 K2 K3
    python find_consultancy_contacts.py --hunter-keys K1 K2

Incremental mode (default):
  Re-running the script will:
    - Load contacts from the existing Excel file
    - Skip searches that were already completed (via .search_cache.json)
    - Skip enrichment for people who already have an email
    - Merge new + old contacts and write a fresh Excel
  This lets you add more keys and re-run without wasting credits.

Fresh start:
  python find_consultancy_contacts.py --fresh   # clears cache + ignores Excel

Where to get keys:
  Apollo.io: https://app.apollo.io/#/settings/integrations/api
  Hunter.io: https://hunter.io/api-keys
        """,
    )
    parser.add_argument("--apollo-keys", nargs="+", metavar="KEY",
                        help="One or more Apollo.io API keys (space-separated)")
    parser.add_argument("--hunter-keys", nargs="+", metavar="KEY",
                        help="One or more Hunter.io API keys (space-separated)")
    parser.add_argument("--no-enrich", action="store_true",
                        help="Skip Apollo enrichment (saves credits, no emails from Apollo)")
    parser.add_argument("--it-only", action="store_true",
                        help="Hunter: only search IT dept (1 credit/company instead of 3)")
    parser.add_argument("--fresh", action="store_true",
                        help="Clear cache and ignore existing Excel (full fresh run)")
    parser.add_argument("--output", default="Bay_Area_IT_Consultancies_Contacts.xlsx",
                        help="Output Excel file path (default: Bay_Area_IT_Consultancies_Contacts.xlsx)")
    # Backwards compatibility
    parser.add_argument("--apollo-key", help=argparse.SUPPRESS)
    parser.add_argument("--hunter-key", help=argparse.SUPPRESS)
    args = parser.parse_args()

    # Load .env file
    load_env_file()

    # Collect Apollo keys
    apollo_keys = []
    if args.apollo_keys:
        apollo_keys = args.apollo_keys
    elif args.apollo_key:
        apollo_keys = [args.apollo_key]
    else:
        apollo_keys = collect_keys_from_env("APOLLO_API_KEY")

    # Collect Hunter keys
    hunter_keys = []
    if args.hunter_keys:
        hunter_keys = args.hunter_keys
    elif args.hunter_key:
        hunter_keys = [args.hunter_key]
    else:
        hunter_keys = collect_keys_from_env("HUNTER_API_KEY")

    if not apollo_keys and not hunter_keys:
        print("=" * 60)
        print("  No API keys provided!")
        print("=" * 60)
        print()
        print("You need at least one API key to search for contacts.")
        print()
        print("Option 1: Create a .env file in this directory:")
        print('  APOLLO_API_KEY_1=your_first_key')
        print('  APOLLO_API_KEY_2=your_second_key')
        print('  HUNTER_API_KEY_1=your_first_key')
        print('  HUNTER_API_KEY_2=your_second_key')
        print()
        print("Option 2: Pass key(s) as arguments:")
        print("  python find_consultancy_contacts.py --apollo-keys KEY1 KEY2")
        print("  python find_consultancy_contacts.py --hunter-keys KEY1 KEY2")
        print()
        print("Where to get keys:")
        print("  Apollo.io: https://app.apollo.io/#/settings/integrations/api")
        print("  Hunter.io: https://hunter.io/api-keys")
        print()
        sys.exit(1)

    # Build key pools
    apollo_pool = KeyPool("Apollo", apollo_keys) if apollo_keys else None
    hunter_pool = KeyPool("Hunter", hunter_keys) if hunter_keys else None

    # Search cache
    cache = SearchCache(".search_cache.json")

    # Load existing contacts from previous Excel (incremental mode)
    existing_contacts = []
    seen = set()
    enriched = set()

    if args.fresh:
        cache.clear()
        print("  [Fresh mode] Cache cleared, starting from scratch")
    else:
        existing_contacts, seen, enriched = load_existing_contacts(args.output)
        if existing_contacts:
            with_email = len([c for c in existing_contacts if c["email"]])
            print(f"  [Resume] Loaded {len(existing_contacts)} existing contacts "
                  f"({with_email} with email) from {args.output}")
            print(f"  [Resume] {cache.count} completed searches in cache")
        else:
            print(f"  [Resume] No existing Excel found — fresh run")

    print()
    print("=" * 60)
    print("  Bay Area IT/CS Consultancy Contact Finder")
    print("=" * 60)
    if apollo_pool:
        print(f"  Apollo.io:  {apollo_pool.total} key(s) loaded")
    else:
        print(f"  Apollo.io:  Not configured")
    if hunter_pool:
        print(f"  Hunter.io:  {hunter_pool.total} key(s) loaded")
    else:
        print(f"  Hunter.io:  Not configured")
    enrich_enabled = apollo_pool and not args.no_enrich
    print(f"  Enrichment: {'Enabled (uses credits)' if enrich_enabled else 'Disabled'}")
    print(f"  Failover:   Enabled (rotates keys on rate limit)")
    print(f"  Incremental:{'OFF (--fresh)' if args.fresh else 'ON (skips existing entries)'}")
    print(f"  Output:     {args.output}")
    print(f"  Firms:      {', '.join(f['name'] for f in CONSULTANCIES)}")
    print("=" * 60)

    contacts = collect_contacts(
        apollo_pool=apollo_pool,
        hunter_pool=hunter_pool,
        enrich_apollo=not args.no_enrich,
        existing_contacts=existing_contacts,
        seen=seen,
        enriched=enriched,
        cache=cache,
        hunter_it_only=args.it_only,
    )

    total_with_email = len([c for c in contacts if c["email"]])
    print(f"\n{'='*60}")
    print(f"  Total contacts (all runs): {len(contacts)}")
    print(f"  With email (all runs):     {total_with_email}")

    if apollo_pool:
        print(f"  Apollo keys: {len(apollo_pool.exhausted)} exhausted, "
              f"{apollo_pool.active_count} still active (of {apollo_pool.total})")
    if hunter_pool:
        print(f"  Hunter keys: {len(hunter_pool.exhausted)} exhausted, "
              f"{hunter_pool.active_count} still active (of {hunter_pool.total})")
    print(f"  Cached searches: {cache.count}")
    print(f"{'='*60}")

    for firm in CONSULTANCIES:
        firm_contacts = [c for c in contacts if c["consultancy"] == firm["name"]]
        with_email = len([c for c in firm_contacts if c["email"]])
        print(f"  {firm['name']:25s} -> {len(firm_contacts):3d} contacts ({with_email} with email)")

    if contacts:
        write_excel(contacts, args.output)
    else:
        print("\nNo contacts found. Check your API keys and try again.")

    print("\nDone! Re-run anytime with new keys — existing results are preserved.")


if __name__ == "__main__":
    main()
