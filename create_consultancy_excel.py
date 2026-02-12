import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()

# ── Sheet 1: Top 5 Consultancies Overview ──
ws1 = wb.active
ws1.title = "Consultancies Overview"

# Styles
header_font = Font(name="Calibri", bold=True, size=12, color="FFFFFF")
header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
subheader_font = Font(name="Calibri", bold=True, size=11)
normal_font = Font(name="Calibri", size=11)
thin_border = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

headers1 = [
    "Rank", "Consultancy Name", "Headquarters / Bay Area Office",
    "Focus Areas (IT/CS)", "Careers Page URL", "Notes"
]

for col_idx, header in enumerate(headers1, 1):
    cell = ws1.cell(row=1, column=col_idx, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = thin_border

consultancies = [
    [
        1,
        "Accenture",
        "San Francisco & Mountain View, CA",
        "Technology consulting, AI/ML, Cloud, Software Engineering, Data Engineering, IT Architecture",
        "https://www.accenture.com/us-en/careers",
        "185+ open roles in Bay Area; mostly hybrid; strong in AI research and technology architecture"
    ],
    [
        2,
        "Deloitte",
        "San Francisco, San Jose, Palo Alto, CA",
        "IT strategy, Technology consulting, Cybersecurity, Cloud, AI/Analytics, Software Development",
        "https://www2.deloitte.com/us/en/careers.html",
        "#1 consulting firm worldwide by revenue (Gartner); launched Zora AI platform; 2000+ tech roles nationally"
    ],
    [
        3,
        "McKinsey & Company",
        "555 California St, San Francisco & Silicon Valley",
        "Digital/Technology transformation, AI, Data analytics, IT strategy, Software engineering",
        "https://www.mckinsey.com/careers",
        "Serves tech, financial services, healthcare in Bay Area; strong digital practice"
    ],
    [
        4,
        "Slalom Consulting",
        "San Francisco & Walnut Creek, CA",
        "Salesforce, Cloud/DevOps, Data architecture, AI/ML engineering, Platform engineering",
        "https://www.slalom.com/us/en/careers",
        "37 open roles in SF area; mid-senior to director level; strong local Bay Area presence"
    ],
    [
        5,
        "Cognizant",
        "San Francisco Bay Area, CA",
        "IT consulting, Digital engineering, Cloud infrastructure, Application modernization, Data & AI",
        "https://careers.cognizant.com",
        "Large global IT services firm; 64+ recruiter positions nationally; significant Bay Area operations"
    ],
]

for row_idx, row_data in enumerate(consultancies, 2):
    for col_idx, value in enumerate(row_data, 1):
        cell = ws1.cell(row=row_idx, column=col_idx, value=value)
        cell.font = normal_font
        cell.alignment = Alignment(vertical="top", wrap_text=True)
        cell.border = thin_border

# Column widths for sheet 1
col_widths1 = [6, 22, 35, 55, 45, 60]
for i, w in enumerate(col_widths1, 1):
    ws1.column_dimensions[get_column_letter(i)].width = w

# ── Sheet 2: IT/CS Recruiting Contacts ──
ws2 = wb.create_sheet("IT CS Recruiting Contacts")

headers2 = [
    "Consultancy", "First Name", "Last Name", "Email",
    "Job Title", "LinkedIn Profile URL", "Phone", "Source", "Notes"
]

for col_idx, header in enumerate(headers2, 1):
    cell = ws2.cell(row=1, column=col_idx, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = thin_border

# Pre-populate consultancy names so the user can fill in contacts
sample_contacts = [
    # These are template rows — company names pre-filled, rest to be completed
    ["Accenture", "", "", "", "IT Recruiter / Talent Acquisition", "", "", "", "Search LinkedIn: 'Accenture recruiter IT San Francisco'"],
    ["Accenture", "", "", "", "Technology Hiring Manager", "", "", "", ""],
    ["Deloitte", "", "", "", "IT Recruiter / Talent Acquisition", "", "", "", "Search LinkedIn: 'Deloitte recruiter technology San Francisco'"],
    ["Deloitte", "", "", "", "Technology Hiring Manager", "", "", "", ""],
    ["McKinsey & Company", "", "", "", "IT Recruiter / Talent Acquisition", "", "", "", "Search LinkedIn: 'McKinsey recruiter digital San Francisco'"],
    ["McKinsey & Company", "", "", "", "Technology Hiring Manager", "", "", "", ""],
    ["Slalom Consulting", "", "", "", "IT Recruiter / Talent Acquisition", "", "", "", "Search LinkedIn: 'Slalom recruiter technology San Francisco'"],
    ["Slalom Consulting", "", "", "", "Technology Hiring Manager", "", "", "", ""],
    ["Cognizant", "", "", "", "IT Recruiter / Talent Acquisition", "", "", "", "Search LinkedIn: 'Cognizant recruiter IT San Francisco'"],
    ["Cognizant", "", "", "", "Technology Hiring Manager", "", "", "", ""],
]

for row_idx, row_data in enumerate(sample_contacts, 2):
    for col_idx, value in enumerate(row_data, 1):
        cell = ws2.cell(row=row_idx, column=col_idx, value=value)
        cell.font = normal_font
        cell.alignment = Alignment(vertical="top", wrap_text=True)
        cell.border = thin_border

# Column widths for sheet 2
col_widths2 = [22, 15, 15, 30, 30, 40, 16, 15, 50]
for i, w in enumerate(col_widths2, 1):
    ws2.column_dimensions[get_column_letter(i)].width = w

# ── Sheet 3: How to Find Contacts ──
ws3 = wb.create_sheet("How to Find Contacts")

tips = [
    ["Tool / Method", "Description", "URL"],
    ["LinkedIn Sales Navigator", "Best for finding recruiters by company, title, and location. Search for 'Recruiter' or 'Talent Acquisition' at each firm in SF Bay Area.", "https://www.linkedin.com/sales"],
    ["Hunter.io", "Find email addresses by company domain. Enter the company domain (e.g., accenture.com) to find publicly indexed emails.", "https://hunter.io"],
    ["Apollo.io", "Free tier available. Search by company, job title, and location to find recruiter contacts with verified emails.", "https://www.apollo.io"],
    ["RocketReach", "Find professional email addresses and phone numbers for people at specific companies.", "https://rocketreach.co"],
    ["LinkedIn Search", "Free method: Search '[Company] recruiter IT San Francisco' on LinkedIn. Send connection requests with a personalized note.", "https://www.linkedin.com"],
    ["Company Careers Page", "Visit each firm's careers page (listed in Sheet 1). Many have 'Contact a Recruiter' options or team pages.", "See Sheet 1"],
    ["Glassdoor", "Check company reviews and sometimes find recruiter names mentioned in interview experience reviews.", "https://www.glassdoor.com"],
    ["Email Pattern Guessing", "Most large firms use patterns like firstname.lastname@company.com. Verify with Hunter.io or similar tools.", ""],
]

for row_idx, row_data in enumerate(tips):
    for col_idx, value in enumerate(row_data, 1):
        cell = ws3.cell(row=row_idx + 1, column=col_idx, value=value)
        if row_idx == 0:
            cell.font = header_font
            cell.fill = header_fill
        else:
            cell.font = normal_font
        cell.alignment = Alignment(vertical="top", wrap_text=True)
        cell.border = thin_border

ws3.column_dimensions["A"].width = 28
ws3.column_dimensions["B"].width = 80
ws3.column_dimensions["C"].width = 40

# Freeze top rows
ws1.freeze_panes = "A2"
ws2.freeze_panes = "A2"
ws3.freeze_panes = "A2"

# Save
output_path = "/Users/ldmello/consult/Bay_Area_IT_Consultancies_Contacts.xlsx"
wb.save(output_path)
print(f"Excel file saved to: {output_path}")
